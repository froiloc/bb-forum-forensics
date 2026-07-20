# =============================================================================
# tests/test_management_export_excel.py
# IT-Forensisches Ermittlungswerkzeug — Baustelle 7: Export-Subsystem (AP-2B)
# =============================================================================
# Testsuite fuer Build 441: Fallstatus -> Excel (build_case_status_xlsx).
#
# XL01 — Arbeitsmappe ladbar; Blatt 'Fallstatus'; Aktenkopf vorhanden
# XL02 — Tabellenkopf = CASE_STATUS_COLUMNS-Kopftexte (Reihenfolge)
# XL03 — Datenzeile korrekt (username, status, Ampel), Zeitstempel formatiert
# XL04 — has_note -> 'ja'/'nein'; None-Werte -> leere Zelle (kein Absturz, GR1)
# XL05 — Erzeugungsvermerk (Ersteller/Build) + Pruefsumme im Fuss
# XL06 — case_rows_digest deterministisch + unabhaengig nachrechenbar
# XL07 — Digest reagiert auf Datenaenderung (Reihenfolge/Wert signifikant)
# XL08 — leere Fallliste: Kopf/Fuss vorhanden, keine Datenzeile, Digest stabil
# XL09 — ExcelUnavailable, wenn openpyxl fehlt (Import-Guard)
# XL10 — UTF-8: multilingualer Benutzername bleibt erhalten
#
# Build 469: Schluesselumstellung user_id -> subject_id (M019)
# Version: v0.7.469 · Build: 469 · 2026-07-20
# =============================================================================

import builtins
import hashlib
import os
import sys
from io import BytesIO

import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from management.export.excel_case_status import (   # noqa: E402
    build_case_status_xlsx,
    case_rows_digest,
    CASE_STATUS_COLUMNS,
    ExcelUnavailable,
    _fmt_ts,
)
from management.export.export_envelope import ExportContext  # noqa: E402
from management.export.checksum import canonical_rows_sha256  # noqa: E402

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook  # noqa: E402


def _ctx(**over):
    base = dict(
        behoerde="Polizei NRW — EK Zarewitsch",
        aktenzeichen="Alle Faelle",
        ersteller="h012345",
        build_number=441,
        generated_at="2026-07-19 12:00:00",
        chain_ok=True,
        chain_tip_seq=42,
        chain_tip_hash="ab" * 32,
    )
    base.update(over)
    return ExportContext(**base)


def _sample_rows():
    return [
        {
            "subject_id": 4711, "username": "täter_süd", "status": "in_progress",
            "priority": 1, "assigned_system_username": "h012345",
            "assigned_display_name": "Muster, Erika", "has_note": 1,
            "approved_at": None, "total_pages_scraped": 87, "event_count": 5,
            "last_event_kind": "CASE_ASSIGNED", "last_event_at": 1_700_000_000,
            "support_count": 0, "last_activity_at": 1_700_000_500,
            "ampel": "gelb", "ampel_reason": "mittlere Inaktivitaet",
            "created_at": 1_699_000_000, "updated_at": 1_700_000_500,
        },
        {
            "subject_id": 4712, "username": "user_中文", "status": "open",
            "priority": 3, "assigned_system_username": None,
            "assigned_display_name": None, "has_note": 0,
            "approved_at": None, "total_pages_scraped": None, "event_count": 0,
            "last_event_kind": None, "last_event_at": None,
            "support_count": 0, "last_activity_at": 1_699_500_000,
            "ampel": "rot", "ampel_reason": "lange inaktiv",
            "created_at": 1_699_000_000, "updated_at": 1_699_500_000,
        },
    ]


def _load(data: bytes):
    return load_workbook(BytesIO(data)).active


def test_xl01_workbook_loads_with_head():
    ws = _load(build_case_status_xlsx(_sample_rows(), _ctx()))
    assert ws.title == "Fallstatus"
    joined = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
    assert "VERTRAULICH" in joined
    assert "Fallstatus-Uebersicht" in joined
    assert "EK Zarewitsch" in joined


def test_xl02_table_header_matches_columns():
    ws = _load(build_case_status_xlsx(_sample_rows(), _ctx()))
    heads = [h for (_k, h, _f) in CASE_STATUS_COLUMNS]
    # Suche die Kopfzeile: erste Zeile, die den ersten Kopftext traegt.
    found = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == heads[0]:
            found = [v for v in row[:len(heads)]]
            break
    assert found == heads


def test_xl03_data_row_values():
    ws = _load(build_case_status_xlsx(_sample_rows(), _ctx()))
    cells = [str(c.value) for col in ws.iter_cols() for c in col if c.value is not None]
    text = "\n".join(cells)
    assert "täter_süd" in text and "in_progress" in text and "gelb" in text
    assert _fmt_ts(1_700_000_000) in text     # Zeitstempel formatiert


def test_xl04_has_note_and_none():
    ws = _load(build_case_status_xlsx(_sample_rows(), _ctx()))
    text = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value is not None)
    assert "ja" in text and "nein" in text


def test_xl05_footer_vermerk_and_checksum():
    rows = _sample_rows()
    ws = _load(build_case_status_xlsx(rows, _ctx()))
    text = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value is not None)
    assert "Erstellt von: Muster, Erika (h012345)".split(":")[0] in text
    assert "h012345" in text
    assert "Werkzeug-Build: 441" in text
    assert case_rows_digest(rows) in text


def test_xl06_digest_independently_recomputable():
    rows = _sample_rows()
    keys = [k for (k, _h, _f) in CASE_STATUS_COLUMNS]
    tables = [("cases", [r.get(k) for k in keys]) for r in rows]
    assert case_rows_digest(rows) == canonical_rows_sha256(tables)


def test_xl07_digest_sensitive_to_change():
    rows = _sample_rows()
    d1 = case_rows_digest(rows)
    rows2 = _sample_rows()
    rows2[0]["status"] = "closed"
    assert case_rows_digest(rows2) != d1
    # Reihenfolge signifikant
    assert case_rows_digest(list(reversed(_sample_rows()))) != d1


def test_xl08_empty_case_list():
    data = build_case_status_xlsx([], _ctx())
    ws = _load(data)
    text = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value is not None)
    assert "VERTRAULICH" in text                      # Kopf da
    assert "Pruefsumme Datensatz (SHA-256):" in text  # Fuss da
    assert case_rows_digest([]) in text               # Digest ueber leere Menge


def test_xl09_excel_unavailable(monkeypatch):
    real_import = builtins.__import__

    def fake_import(name, *a, **k):
        if name == "openpyxl":
            raise ImportError("simuliert fehlend")
        return real_import(name, *a, **k)

    monkeypatch.setattr(builtins, "__import__", fake_import)
    with pytest.raises(ExcelUnavailable):
        build_case_status_xlsx(_sample_rows(), _ctx())


def test_xl10_utf8_preserved():
    ws = _load(build_case_status_xlsx(_sample_rows(), _ctx()))
    text = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value is not None)
    assert "user_中文" in text
