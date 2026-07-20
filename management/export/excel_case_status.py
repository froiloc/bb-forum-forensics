# =============================================================================
# management/export/excel_case_status.py
# IT-Forensisches Ermittlungswerkzeug — Baustelle 7: Export-Subsystem (AP-2B)
# =============================================================================
# Zweck (Idee 2 — Fallstatus -> Excel):
#   Reiner Renderer, der eine Liste von Fall-Uebersichten (CaseOverview via
#   dataclasses.asdict, Beleg management/dashboard/dashboard_repo.py:206) in eine
#   EINZELNE .xlsx-Arbeitsmappe giesst. Der Rahmen (Aktenkopf,
#   Erzeugungsvermerk, Pruefsumme) kommt aus dem gemeinsamen Export-Framework
#   (Build 440, management.export.ExportEnvelope) — so ist der Excel-Export
#   gleich gerahmt wie kuenftige Akten-/StA-Exporte.
#
#   PRUEFSUMME: bewusst ueber die ROHEN Datenspalten (canonical_rows_sha256,
#   deckungsgleich zur Sealer-Kanonik), NICHT ueber die .xlsx-Bytes. Grund:
#   openpyxl bettet Zeitstempel in die Zip-Struktur ein -> Datei-Bytes waeren
#   nicht deterministisch. Der Datendigest hingegen ist stabil und vom Empfaenger
#   unabhaengig aus den Rohwerten nachrechenbar (Ueberpruefbarkeit = oberste
#   Fallregel).
#
#   REINE FUNKTION: kein DB-/Netz-/Uhr-Zugriff; alle Werte (Daten + Kontext)
#   werden injiziert -> vollstaendig automatisiert testbar (Muster der
#   bestehenden html_export.py). openpyxl wird LAZY importiert; fehlt sie in der
#   Offline-VM -> ExcelUnavailable (klarer 503-artiger Fehler, analog
#   DocxRendererUnavailable/PdfRendererUnavailable), kein stiller Ausfall.
#
# Version: v0.7.469 · Build: 469 · 2026-07-20
# =============================================================================

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import List, Optional, Sequence, Tuple

from management.export.checksum import canonical_rows_sha256
from management.export.export_envelope import ExportContext, ExportEnvelope


class ExcelUnavailable(RuntimeError):
    """openpyxl ist in der Laufzeitumgebung nicht installiert."""


# Spaltenordnung des Blatts: (dict-Schluessel aus CaseOverview, Kopftext,
# Formatierung). 'ts' -> UTC-Zeitstempel, sonst Rohwert. Die REIHENFOLGE ist
# zugleich die Reihenfolge fuer die Pruefsumme (Rohwerte) -> stabil belegbar.
CASE_STATUS_COLUMNS: Sequence[Tuple[str, str, str]] = (
    ("subject_id",               "Subject-ID",         "raw"),
    ("username",                 "Benutzername",       "raw"),
    ("status",                   "Fallstatus",         "raw"),
    ("priority",                 "Prioritaet",         "raw"),
    ("assigned_system_username", "Zugewiesen (Konto)", "raw"),
    ("assigned_display_name",    "Zugewiesen (Name)",  "raw"),
    ("has_note",                 "Notiz",              "raw"),
    ("approved_at",              "Freigegeben am",     "ts"),
    ("total_pages_scraped",      "Seiten gescrapt",    "raw"),
    ("event_count",              "Ereignisse",         "raw"),
    ("last_event_kind",          "Letzte Ereignisart", "raw"),
    ("last_event_at",            "Letztes Ereignis",   "ts"),
    ("support_count",            "Support aktiv",      "raw"),
    ("last_activity_at",         "Letzte Aktivitaet",  "ts"),
    ("ampel",                    "Ampel",              "raw"),
    ("ampel_reason",             "Ampel-Begruendung",  "raw"),
    ("created_at",               "Angelegt am",        "ts"),
    ("updated_at",               "Aktualisiert am",    "ts"),
)


def _fmt_ts(ts) -> str:
    """UTC-Zeitstempel als String; None -> '-'."""
    if ts is None:
        return "-"
    return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime(
        "%Y-%m-%d %H:%MZ"
    )


def case_rows_digest(rows: List[dict]) -> str:
    """
    Datendigest ueber die ROHEN Spaltenwerte in CASE_STATUS_COLUMNS-Ordnung.
    Deterministisch und unabhaengig nachrechenbar (canonical_rows_sha256).
    Fehlende Schluessel -> None (kein stiller Ausfall; die Spalte existiert,
    ist aber leer belegt).
    """
    keys = [k for (k, _h, _f) in CASE_STATUS_COLUMNS]
    tables = [("cases", [row.get(k) for k in keys]) for row in rows]
    return canonical_rows_sha256(tables)


def build_case_status_xlsx(
    rows: List[dict],
    context: ExportContext,
    *,
    scope_label: Optional[str] = None,
) -> bytes:
    """
    Baut die .xlsx-Arbeitsmappe und gibt ihre Bytes zurueck.

    rows        — Liste von CaseOverview-dicts (dataclasses.asdict).
    context     — ExportContext (Behoerde/Aktenzeichen/Ersteller/Build/Zeit/
                  Kettenspitze) fuer den einheitlichen Rahmen.
    scope_label — optionaler Umfangs-Vermerk (z. B. 'Alle Faelle'), wird in den
                  Aktenkopf uebernommen.

    Fehlt openpyxl -> ExcelUnavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:  # pragma: no cover - Umgebungsrandfall
        raise ExcelUnavailable(
            "openpyxl ist nicht installiert — Excel-Export nicht moeglich."
        ) from exc

    env = ExportEnvelope(context)
    digest = case_rows_digest(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fallstatus"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF")

    r = 1
    # -- Aktenkopf (Klassifikation, Titel, Behoerde/Aktenzeichen) ------------
    titel = "Fallstatus-Uebersicht"
    if scope_label:
        titel += " — " + scope_label
    ws.cell(row=r, column=1, value=context.klassifikation).font = bold
    r += 1
    ws.cell(row=r, column=1, value=titel).font = Font(bold=True, size=14)
    r += 1
    ws.cell(row=r, column=1,
            value="Behoerde: %s · Aktenzeichen: %s"
                  % (context.behoerde, context.aktenzeichen))
    r += 2  # Leerzeile

    # -- Tabellenkopf --------------------------------------------------------
    header_row = r
    for c, (_key, head, _fmt) in enumerate(CASE_STATUS_COLUMNS, start=1):
        cell = ws.cell(row=r, column=c, value=head)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    r += 1

    # -- Datenzeilen ---------------------------------------------------------
    for row in rows:
        for c, (key, _head, fmt) in enumerate(CASE_STATUS_COLUMNS, start=1):
            val = row.get(key)
            if fmt == "ts":
                out = _fmt_ts(val)
            elif key == "has_note":
                out = "ja" if val else "nein"
            else:
                out = "" if val is None else val
            ws.cell(row=r, column=c, value=out)
        r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # -- Erzeugungsvermerk + Pruefsumme (Fuss) -------------------------------
    r += 1
    ws.cell(row=r, column=1, value="Erzeugungsvermerk").font = bold
    r += 1
    for line in env.erzeugungsvermerk_lines():
        ws.cell(row=r, column=1, value=line)
        r += 1
    ws.cell(row=r, column=1,
            value="Pruefsumme Datensatz (SHA-256): %s" % digest).font = bold

    # -- Spaltenbreiten (Lesbarkeit) -----------------------------------------
    for c, (_key, head, _fmt) in enumerate(CASE_STATUS_COLUMNS, start=1):
        letter = ws.cell(row=header_row, column=c).column_letter
        ws.column_dimensions[letter].width = max(12, min(28, len(head) + 4))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
